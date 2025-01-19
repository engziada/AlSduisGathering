import io
import os
import random
import secrets
import shutil
from io import BytesIO

import arabic_reshaper
import pandas as pd
import requests
from bidi.algorithm import get_display
from bs4 import BeautifulSoup
from datetime import datetime
from flask import (
    Flask,
    Response,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_cors import CORS
from flask_migrate import Migrate
from flask_sqlalchemy import SQLAlchemy
from flask_wtf import FlaskForm
from icecream import ic
from PIL import Image, ImageDraw, ImageFont, __version__, features
from openpyxl import Workbook
from sqlalchemy import and_, or_
from werkzeug.utils import secure_filename
from wtforms import (
    HiddenField,
    IntegerField,
    SelectField,
    SelectMultipleField,
    StringField,
    SubmitField,
    TextAreaField,
    ValidationError,
    validators,
    widgets,
)
from wtforms.validators import DataRequired
import json
from functools import wraps

# ==============================================================================

# Initialize Flask
app = Flask(__name__)

# Set database path based on environment
if os.environ.get('RENDER'):
    # Use Render's persistent storage
    db_path = '/var/data/registrations.db'
    os.makedirs('/var/data', exist_ok=True)
else:
    # Local development - use instance folder
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance', 'registrations.db')
    os.makedirs(os.path.dirname(db_path), exist_ok=True)

app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{db_path}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["Close_Registrations"] = False
app.secret_key = secrets.token_hex(16)  

# Initialize Flask-SQLAlchemy
db = SQLAlchemy(app)
migrate = Migrate(app, db)  # Initialize Flask-Migrate

# Initialize Flask-CORS
CORS(app)

# Google spreadsheet details
SPREADSHEET_URL = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vRaFYu_tsagGZ16_B9ku1LurJxfe4JtN-6vW5a1_cAEEWkVsV7Wy9hm0lttiYTeBCnjOmnHJjTV6MNd/pubhtml'
GUESTS_SHEET_NAME = 'Sheet1'

# Passcode whitelist
PASSCODE_WHITELIST = ['9753', '6290']
RECEPTIONIST_WHITELIST = ['1234']  # Add more receptionist codes as needed

# ==============================================================================
# ''' Models '''
# ==============================================================================


# Define the Registration model for the database table
class Registration(db.Model):
    __tablename__ = "registration"
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    phone_number = db.Column(db.String(16), unique=True, nullable=False)
    first_name = db.Column(db.Text, nullable=False)
    family_name = db.Column(db.Text, nullable=False)
    father_name = db.Column(db.Text)
    first_grand_name = db.Column(db.Text)
    second_grand_name = db.Column(db.Text)
    third_grand_name = db.Column(db.Text)
    relation = db.Column(db.Text)
    age = db.Column(db.String(3))
    gender = db.Column(db.String(15))
    city = db.Column(db.Text)
    attendance = db.Column(db.Text, nullable=False)
    ideas = db.Column(db.Text)
    registration_number = db.Column(db.String(4), unique=True)
    prize_id = db.Column(db.Integer, db.ForeignKey("prize.id", name="fk_reg_prize_id"))
    is_attended = db.Column(db.Boolean, default=False)
    # Add relationship to Prize model
    prize = db.relationship('Prize', backref='winner', lazy=True)


# Define the Prize model
class Prize(db.Model):
    __tablename__ = "prize"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    guest_registration_number = db.Column(db.String(4))
    allowed_families = db.Column(db.String(100))
    allowed_age_range = db.Column(db.String(100))
    allowed_gender = db.Column(db.String(100))
    is_next = db.Column(db.Boolean, default=False)

    def __init__(
        self,
        name,
        description=None,
        guest_registration_number=None,
        allowed_families=None,
        allowed_age_range=None,
        allowed_gender=None,
        is_next=False,
    ):
        self.name = name
        self.description = description
        self.guest_registration_number = guest_registration_number
        self.allowed_families = allowed_families
        self.allowed_age_range = allowed_age_range
        self.allowed_gender = allowed_gender
        self.is_next = is_next


# Define the Children model for the database table
class Children(db.Model):
    __tablename__ = "children"
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    parent_phone = db.Column(db.String(16), db.ForeignKey('registration.phone_number'), nullable=False)
    first_name = db.Column(db.Text, nullable=False)
    father_name = db.Column(db.Text, nullable=False)
    grandfather_name = db.Column(db.Text, nullable=False)
    family_name = db.Column(db.Text, nullable=False)
    gender = db.Column(db.String(15), nullable=False)
    age = db.Column(db.Integer, nullable=False)
    emergency_phone = db.Column(db.String(16), nullable=False)  # Adding emergency phone field
    registration_number = db.Column(db.String(4))  # Same as parent's

    def __repr__(self):
        return f'<Child {self.first_name} {self.father_name}>'

# ==============================================================================
#   ''' Database and Tables '''
# ==============================================================================

# Create the database and tables if they don't exist

# Initialize Flask-Faker
# from faker import Faker
# fake = Faker()

with app.app_context():
    db.create_all()

    # Loop to create 50 fake records
    # for _ in range(50):
    #     record = Registration(
    #         phone_number=fake.phone_number(),
    #         first_name=fake.first_name(),
    #         family_name=fake.last_name(),
    #         father_name=fake.last_name(),
    #         first_grand_name=fake.last_name(),
    #         second_grand_name=fake.last_name(),
    #         third_grand_name=fake.last_name(),
    #         relation=fake.word(),
    #         age=str(fake.random_int(18, 80)),  # Assuming age between 18 and 80
    #         gender=fake.random_element(["Male", "Female"]),
    #         city=fake.city(),
    #         attendance=fake.random_element(["Yes", "No"]),
    #         ideas=fake.text(),
    #         registration_number=str(fake.random_int(100, 999))  # Unique 3-digit registration number
    #         # Add prize_id if needed
    #     )
    #     db.session.add(record)

    # Commit the changes
    db.session.commit()

# ==============================================================================
# ''' Helper Functions '''
# ==============================================================================

def validate_non_family_name(form, field):
    if form.family_name.data == "أخرى" and not field.data:
        raise ValidationError("من فضلك أدخل البيان المطلوب")

# Helper function to generate a unique registration number
def generate_registration_number():
    return str(random.randint(1000, 9999))


def is_registered(phone_number):
    return Registration.query.filter_by(phone_number=phone_number).first() is not None


def get_user_data(phone_number):
    if not phone_number:
        users = Registration.query.all()
        users_data = []
        for user in users:
            user_data = {
                "phone_number": user.phone_number,
                "first_name": user.first_name,
                "family_name": user.family_name,
                "father_name": user.father_name,
                "first_grand_name": user.first_grand_name,
                "second_grand_name": user.second_grand_name,
                "third_grand_name": user.third_grand_name,
                "relation": user.relation,
                "age": user.age,
                "gender": user.gender,
                "city": user.city,
                "attendance": user.attendance,
                "ideas": user.ideas,
                "registration_number": user.registration_number,
            }
            users_data.append(user_data)
        return users_data
    else:
        user = Registration.query.filter_by(phone_number=phone_number).first()
        if user:
            return {
                "phone_number": user.phone_number,
                "first_name": user.first_name,
                "family_name": user.family_name,
                "father_name": user.father_name,
                "first_grand_name": user.first_grand_name,
                "second_grand_name": user.second_grand_name,
                "third_grand_name": user.third_grand_name,
                "relation": user.relation,
                "age": user.age,
                "gender": user.gender,
                "city": user.city,
                "attendance": user.attendance,
                "ideas": user.ideas,
                "registration_number": user.registration_number,
            }
        else:
            return {}


def create_image(content):
    # Create an image with a white background
    width, height = 800, 600  # Set the image size as needed
    image = Image.new("RGB", (width, height), color="white")
    # Create a drawing context
    draw = ImageDraw.Draw(image)

    # Define additional styles (e.g., colors)
    text_color = "black"
    background_color = "white"
    top_stripe_color = (8, 94, 157)
    bottom_stripe_color = (88, 88, 86)
    font_size = 36

    static_dir = os.path.join(os.path.dirname(__file__), "static")
    font_file_path = os.path.join(static_dir, "fonts", "Tahoma.ttf")
    # Set the font and font size (you may need to download and specify an Arabic font)
    font = ImageFont.truetype(font_file_path, size=font_size)

    # Split content into lines
    lines = content.split("\n")

    # Calculate the total height of all lines
    total_height = len(lines) * font_size

    # Calculate the vertical position (centered)
    y = (height - total_height) // 2

    # Draw a background rectangle with a specified color
    draw.rectangle([0, 0, width, height], fill=background_color)

    # Draw the horizontal stripe at the top of the image
    draw.rectangle([0, 0, width, 15], fill=top_stripe_color)
    draw.rectangle([0, height - 15, width, height], fill=bottom_stripe_color)

    # Draw each line with specified styles, centered horizontally
    idx = 0
    for line in lines:
        idx += 1
        print(f"Line no. ({idx}): ", line)
        print("BBox: ", font.getmask(line).getbbox())
        if idx == 8:
            break
        if (
            not line
            or line.isspace()
            or line == "\n"
            or line == "\r"
            or line == "\r\n"
            or line == "\n\r"
        ):
            continue
        
        reshaped_text = arabic_reshaper.reshape(line)
        bidi_text = get_display(reshaped_text)
        
        text_color = "red" if idx in (3, 5, 7) else "black"
        left, top, right, bottom = font.getmask(line).getbbox()
        text_width, text_height = right - left, bottom - top
        x = (
            width - text_width
        ) // 2  # Calculate horizontal position for center alignment
        
        draw.text((x, y), bidi_text, fill=text_color, font=font)
        # draw.text((x, y), line, fill=text_color, font=font)
        y += font_size + 20  # Adjust vertical position for the next line

    # # Set the position to start drawing
    # x, y = 10, 10
    # # Draw the content onto the image
    # draw.text((x, y), content, fill='black', font=font)
    return image

# Function to check if a value exists in a specific sheet
# def is_value_in_sheet(spreadsheet_url, sheet_name, target_value):
#     # Make a request to fetch the sheet data
#     response = requests.get(spreadsheet_url)
#     if response.status_code == 200:
#         # Parse the HTML response
#         soup = BeautifulSoup(response.content, 'html.parser')
#         # Find the table corresponding to the sheet
#         tables = soup.find('table')
#         # Check if the table exists
#         if tables:
#             # Check if the target value exists in any cell
#             for row in tables.find_all('tr'):
#                 if target_value in [cell.text.strip() for cell in row.find_all('td')]:
#                     return True
#             return False
#         else:
#             print(f"Sheet '{sheet_name}' not found.")
#             return False
#     else:
#         print(f"Error fetching data from Spreadsheet. Status code: {response.status_code}")
#         return False

# ==============================================================================
# ''' Forms '''
# ==============================================================================
class MultiCheckboxField(SelectMultipleField):
    widget = widgets.ListWidget(prefix_label=False)
    option_widget = widgets.CheckboxInput()


# Registration form
class RegistrationForm(FlaskForm):
    phone_number = StringField(
        "رقم الهاتف الجوال",
        validators=[
            validators.DataRequired(),
            validators.Regexp(
                "^[+]?\d{10,15}$",
            ),
        ],
    )
    first_name = StringField("اسمك الأول", validators=[validators.InputRequired()])
    family_name = SelectField(
        "اسم العائلة",
        choices=[("السديس", "السديس"), ("أخرى", "أخرى")],
        validators=[validators.InputRequired()],
    )
    custom_family_name = StringField(
        "اسم العائلة غير أسرة السديس", validators=[validate_non_family_name]
    )
    relation = StringField(
        "إن لم تكن من أسرة السديس فلطفاً حدد نوع العلاقة",
        validators=[validate_non_family_name],
    )
    father_name = StringField("اسم الأب", validators=[validators.InputRequired()])
    first_grand_name = StringField(
        "اسم الجد الأول", validators=[validators.InputRequired()]
    )
    second_grand_name = StringField(
        "اسم الجد الثاني", validators=[validators.InputRequired()]
    )
    third_grand_name = StringField(
        "اسم الجد الثالث/فرع الأسرة", validators=[validators.InputRequired()]
    )
    age = SelectField(
        "ماهي فئتك العمرية",
        choices=[
            ("", ""),
            ("أقل من 10 سنوات", "أقل من 10 سنوات"),
            ("من 11 سنة حتى 20 سنة", "من 11 سنة حتى 20 سنة"),
            ("من 21 سنة حتى 40 سنة", "من 21 سنة حتى 40 سنة"),
            ("من 41 سنة حتى 60 سنة", "من 41 سنة حتى 60 سنة"),
            ("من 60 سنة فاكثر", "من 60 سنة فاكثر"),
        ],
        validators=[validators.InputRequired()],
    )
    gender = SelectField(
        "الجنس",
        choices=[
            ("", ""),
            ("ذكر", "ذكر"), 
            ("أنثى", "أنثى")
        ],
        validators=[validators.InputRequired()],
    )
    city = StringField(
        "مدينة العنوان الدائم لك", validators=[validators.InputRequired()]
    )
    attendance = SelectField(
        "هل ستحضر اللقاء",
        choices=[
            ("", ""),
            ("سوف أحضر باذن الله", "سوف أحضر باذن الله"),
            ("أعتذر عن الحضور", "أعتذر عن الحضور"),
        ],
        validators=[validators.InputRequired()],
    )
    ideas = TextAreaField(
        "هل لديك مشاركة أو فكرة تود تقديمها في الحفل ؟ نسعد بمعرفة ذلك"
    )
    children_data = TextAreaField("بيانات الأطفال")
    submit = SubmitField("تسجيل")


# Adding a new prize
class PrizeForm(FlaskForm):
    name = StringField("اسم الهدية", validators=[DataRequired()])
    description = TextAreaField("الوصف")
    allowed_families = StringField("العائلات المسموح لها")
    allowed_age_range = SelectField(
        "العمر المسموح به",
        choices=[
            ("الكل", "الكل"),
            ("أقل من 10 سنوات", "أقل من 10 سنوات"),
            ("من 11 سنة حتى 20 سنة", "من 11 سنة حتى 20 سنة"),
            ("من 21 سنة حتى 40 سنة", "من 21 سنة حتى 40 سنة"),
            ("من 41 سنة حتى 60 سنة", "من 41 سنة حتى 60 سنة"),
            ("من 60 سنة فاكثر", "من 60 سنة فاكثر"),
        ],
    )
    allowed_gender = SelectField(
        "الجنس المسموح به", choices=[("الكل", "الكل"), ("ذكر", "ذكر"), ("أنثى", "أنثى")]
    )
    # guest_registration_number = StringField('Guest Registration Number', validators=[DataRequired()])


# Prize filter form
class FilterForm(FlaskForm):
    family_name = MultiCheckboxField(
        "اسم العائلة", choices=[("السديس", "السديس"), ("أخرى", "أخرى")]
    )
    age = MultiCheckboxField(
        "الفئة العمرية",
        choices=[
            ("أقل من 10 سنوات", "أقل من 10 سنوات"),
            ("من 11 سنة حتى 20 سنة", "من 11 سنة حتى 20 سنة"),
            ("من 21 سنة حتى 40 سنة", "من 21 سنة حتى 40 سنة"),
            ("من 41 سنة حتى 60 سنة", "من 41 سنة حتى 60 سنة"),
            ("من 60 سنة فاكثر", "من 60 سنة فاكثر"),
        ],
    )
    gender = MultiCheckboxField("الجنس", choices=[("ذكر", "ذكر"), ("أنثى", "أنثى")])
    prize_id = IntegerField("الهدية", validators=[DataRequired()])
    submit = SubmitField("بحث")

# Close the registration form
class CloseRegistrationForm(FlaskForm):
    close_registrations = app.config["Close_Registrations"]
    close_status=HiddenField("close_status", default=close_registrations)

# Login form
class LoginForm(FlaskForm):
    passcode = StringField('كود الدخول', validators=[DataRequired()])
    submit = SubmitField('دخول')

# ------------------------------------------------------------------------------
# ''' Registration Routes '''
# ------------------------------------------------------------------------------

# Route for the home page
@app.route("/", methods=["GET", "POST"])
def index():
    close_registrations = app.config["Close_Registrations"]
    
    form = RegistrationForm()  # Create an instance of the RegistrationForm
    if request.method == "POST":
        phone_number = request.form["phone_number"]
        # Check if the number is in the spreadsheet
        # if not is_value_in_sheet(
        #     spreadsheet_url=SPREADSHEET_URL,
        #     sheet_name=GUESTS_SHEET_NAME,
        #     target_value=phone_number if phone_number[0]!='0' else phone_number[1:],
        # ):
    #         # If the number is not in the spreadsheet, redirect to the error page
    #         return render_template("not_in_sheet.html")
        
        # Check if the registration is closed
        if close_registrations:
            return render_template("closed.html")

        # Check if the phone number is not registered
        if not is_registered(phone_number):
            if close_registrations:
                return render_template("closed.html")
            else:
                return redirect(url_for("register", phone_number=phone_number))
        # Check if the phone number is alerady registered
        return redirect(url_for("registered", phone_number=phone_number))
    return render_template("index.html", form=form)


# Route for the registration form
@app.route("/register/<phone_number>", methods=["GET", "POST"])
def register(phone_number):
    # Set the phone number in the form
    form = RegistrationForm()
    form.phone_number.data = phone_number
    
    # Get existing registration if any
    registration = Registration.query.filter_by(phone_number=phone_number).first()
    
    if request.method == "GET" and registration:
        # Pre-fill form with existing data
        form.first_name.data = registration.first_name
        form.family_name.data = registration.family_name
        form.father_name.data = registration.father_name
        form.first_grand_name.data = registration.first_grand_name
        form.second_grand_name.data = registration.second_grand_name
        form.third_grand_name.data = registration.third_grand_name
        form.relation.data = registration.relation
        form.age.data = registration.age
        form.gender.data = registration.gender
        form.city.data = registration.city
        form.attendance.data = registration.attendance
        form.ideas.data = registration.ideas
        
        # Get children data if any
        children = Children.query.filter_by(parent_phone=phone_number).all()
        if children:
            children_data = []
            for child in children:
                child_data = {
                    "first_name": child.first_name,
                    "father_name": child.father_name,
                    "grandfather_name": child.grandfather_name,
                    "family_name": child.family_name,
                    "gender": child.gender,
                    "age": child.age,
                    "emergency_phone": child.emergency_phone
                }
                children_data.append(child_data)
            form.children_data.data = json.dumps(children_data)
    
    if form.validate_on_submit():
        try:
            if not registration:
                registration = Registration()
                registration.registration_number = generate_registration_number()
            
            # Always use the phone number from the URL
            registration.phone_number = phone_number
            registration.first_name = form.first_name.data
            registration.family_name = form.family_name.data
            registration.father_name = form.father_name.data
            registration.first_grand_name = form.first_grand_name.data
            registration.second_grand_name = form.second_grand_name.data
            registration.third_grand_name = form.third_grand_name.data
            registration.relation = form.relation.data
            registration.age = form.age.data
            registration.gender = form.gender.data
            registration.city = form.city.data
            registration.attendance = form.attendance.data
            registration.ideas = form.ideas.data

            if not registration.id:
                db.session.add(registration)
            
            # Handle children data
            if form.children_data.data:
                # Delete existing children first if updating
                if registration.id:
                    Children.query.filter_by(parent_phone=phone_number).delete()
                
                children_data = json.loads(form.children_data.data)
                for child_data in children_data:
                    child = Children(
                        parent_phone=phone_number,  # Always use the phone number from the URL
                        first_name=child_data['first_name'],
                        father_name=child_data['father_name'],
                        grandfather_name=child_data['grandfather_name'],
                        family_name=child_data['family_name'],
                        gender=child_data['gender'],
                        age=int(child_data['age']),
                        emergency_phone=child_data['emergency_phone'],
                        registration_number=registration.registration_number
                    )
                    db.session.add(child)

            db.session.commit()
            return redirect(url_for("registered", phone_number=phone_number))
        except Exception as e:
            db.session.rollback()
            flash(f"حدث خطأ أثناء التسجيل: {str(e)}", "error")
            return render_template("register.html", form=form, registration=registration, phone_number=phone_number)

    return render_template("register.html", form=form, registration=registration, phone_number=phone_number)


@app.route("/registered/<phone_number>", methods=["POST", "GET"])
def registered(phone_number):
    user_data = get_user_data(phone_number)
    if user_data:
        # Get children information
        children = Children.query.filter_by(parent_phone=phone_number).all()
        children_names = [child.first_name for child in children]
        return render_template(
            "registered.html", 
            user_data=user_data, 
            children_names=children_names,
            children_count=len(children)
        )
    return redirect(url_for("index"))

@app.route("/delete/<phone_number>", methods=["POST"])
def delete(phone_number):
    # Verify registration exists
    registration = Registration.query.filter_by(phone_number=phone_number).first()
    if not registration:
        flash("التسجيل غير موجود", "error")
        return redirect(url_for("index"))
    
    try:
        # Delete associated children first
        Children.query.filter_by(parent_phone=phone_number).delete()
        # Delete registration
        db.session.delete(registration)
        db.session.commit()
        flash("تم حذف التسجيل بنجاح", "success")
    except Exception as e:
        db.session.rollback()
        flash("حدث خطأ أثناء حذف التسجيل", "error")
        
    return redirect(url_for("index"))

# Role-based authentication decorator
def role_required(allowed_roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'role' not in session:
                flash('يرجى تسجيل الدخول أولاً', 'error')
                if request.endpoint != 'admin_login':  # Prevent redirect loop
                    return redirect(url_for('admin_login'))
                return f(*args, **kwargs)
            if session['role'] not in allowed_roles:
                flash('غير مصرح لك بالوصول لهذه الصفحة', 'error')
                if session['role'] == 'receptionist':
                    return redirect(url_for('confirm_attendence'))
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Admin authentication decorator (for backward compatibility)
def admin_required(f):
    return role_required(['admin'])(f)

# Admin login route
@app.route("/admin_login", methods=["GET", "POST"])
def admin_login():
    # If already logged in, redirect to appropriate page
    if 'role' in session:
        if session['role'] == 'admin':
            return redirect(url_for('admin'))
        elif session['role'] == 'receptionist':
            return redirect(url_for('confirm_attendence'))
            
    form = LoginForm()
    if form.validate_on_submit():
        passcode = form.passcode.data
        if passcode in PASSCODE_WHITELIST:
            session['role'] = 'admin'
            session.permanent = True  # Make session persistent
            flash('تم تسجيل الدخول بنجاح', 'success')
            return redirect(url_for('admin'))
        elif passcode in RECEPTIONIST_WHITELIST:
            session['role'] = 'receptionist'
            session.permanent = True  # Make session persistent
            flash('تم تسجيل الدخول بنجاح', 'success')
            return redirect(url_for('confirm_attendence'))
        else:
            flash('كود الدخول غير صحيح', 'error')
    return render_template("admin_login.html", form=form)

@app.route("/admin_logout")
def admin_logout():
    session.pop('role', None)
    flash('تم تسجيل الخروج بنجاح', 'success')
    return redirect(url_for('index'))

@app.route("/admin/", methods=["GET"])
@app.route("/admin", methods=["GET"])
@role_required(['admin'])
def admin():
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    # Get guest statistics with prize information
    guests_query = Registration.query.outerjoin(Prize, Registration.prize_id == Prize.id)
    guests = guests_query.order_by(Registration.id.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    guests_count = guests_query.count()
    guest_attendance = guests_query.filter(
        Registration.attendance == "سوف أحضر باذن الله"
    ).count()
    guest_not_attendance = guests_query.filter(
        Registration.attendance == "أعتذر عن الحضور"
    ).count()
    guest_is_attended = guests_query.filter(Registration.is_attended.is_(True)).count()
    guest_male = guests_query.filter(Registration.gender == "ذكر").count()
    guest_female = guests_query.filter(Registration.gender == "أنثى").count()
    
    # Get children statistics
    children_count = Children.query.count()
    children_male = Children.query.filter(Children.gender == "ذكر").count()
    children_female = Children.query.filter(Children.gender == "أنثى").count()
    
    # Create form for registration status
    close_registrations_form = CloseRegistrationForm()
    close_registrations = app.config["Close_Registrations"]
    
    return render_template(
        'admin.html',
        guests=guests,
        guests_count=guests_count,
        guest_attendance=guest_attendance,
        guest_not_attendance=guest_not_attendance,
        guest_is_attended=guest_is_attended,
        guest_male=guest_male,
        guest_female=guest_female,
        children_count=children_count,
        children_male=children_male,
        children_female=children_female,
        form=close_registrations_form,
        close_registrations=close_registrations,
    )

@app.route("/admin/", methods=["POST"])
@app.route("/admin", methods=["POST"])
@role_required(['admin'])
def admin_post():
    close_registrations = app.config["Close_Registrations"]
    close_registrations_form = CloseRegistrationForm()
    
    if (request.method == "POST" and "reg_status_form" in request.form 
        and close_registrations_form.validate_on_submit()):
        app.config["Close_Registrations"] = not close_registrations
        close_registrations_form.close_status.data = not close_registrations
        flash("تم تغيير حالة التسجيل بنجاح", "success")
        return redirect(url_for("admin"))
    
    page = request.args.get("page", 1, type=int)
    per_page = 10
    guests = Registration.query.order_by(Registration.id.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    return render_template('admin.html', guests=guests)

# Delete Guest
@app.route("/delete_guest/<string:guest_phoneno>")
@role_required(['admin'])
def delete_guest(guest_phoneno):
    try:
        # Get the guest record
        guest = Registration.query.filter_by(phone_number=guest_phoneno).first()
        if guest:
            # If guest has a prize, reset the prize's guest_registration_number
            if guest.prize_id:
                prize = Prize.query.get(guest.prize_id)
                if prize:
                    prize.guest_registration_number = None
                    db.session.add(prize)
            
            # Delete the guest
            db.session.delete(guest)
            db.session.commit()
            flash("تم حذف الضيف بنجاح", "success")
        else:
            flash("لم يتم العثور على الضيف", "error")
        return redirect(url_for("admin"))
    except Exception as e:
        db.session.rollback()
        flash(f"حدث خطأ أثناء حذف الضيف: {str(e)}", "error")
        return redirect(url_for("admin"))

@app.route("/export")
@role_required(['admin'])
def export_to_excel():
    # Create a new workbook
    wb = Workbook()
    
    # Create guests sheet
    ws_guests = wb.active
    ws_guests.title = "بيانات الضيوف"
    
    # Add headers for guests
    headers = [
        "الإسم الأول",
        "إسم الأب",
        "إسم الجد الأول",
        "إسم الجد الثاني",
        "إسم الجد الثالث",
        "إسم العائلة",
        "رقم الجوال",
        "رقم التسجيل",
        "صلة القرابة",
        "السن",
        "الجنس",
        "المدينة",
        "تأكيد الحضور",
        "المقترحات",
        "تم الحضور",
        "فاز بجائزة",
        "اسم الجائزة"
    ]
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws_guests.cell(row=1, column=col)
        cell.value = header
        # Apply style
        cell.font = cell.font.copy(bold=True)
        cell.alignment = cell.alignment.copy(horizontal="center")
    
    # Get all guests
    guests = Registration.query.outerjoin(Prize, Registration.prize_id == Prize.id).all()
    
    # Write guest data
    for row, guest in enumerate(guests, start=2):
        data = [
            guest.first_name,
            guest.father_name,
            guest.first_grand_name,
            guest.second_grand_name,
            guest.third_grand_name,
            guest.family_name,
            guest.phone_number,
            guest.registration_number,
            guest.relation or "",
            guest.age,
            guest.gender,
            guest.city,
            "نعم" if guest.attendance == "سوف أحضر باذن الله" else "لا",
            guest.ideas or "",
            "نعم" if guest.is_attended else "لا",
            "نعم" if guest.prize_id else "لا",
            guest.prize.name if guest.prize_id else "-"
        ]
        
        for col, value in enumerate(data, start=1):
            cell = ws_guests.cell(row=row, column=col)
            cell.value = value
            cell.alignment = cell.alignment.copy(horizontal="center")
    
    # Create children sheet
    ws_children = wb.create_sheet(title="بيانات الأطفال")
    
    # Add headers for children
    children_headers = [
        "الإسم الأول",
        "إسم الأب",
        "إسم الجد",
        "إسم العائلة",
        "الجنس",
        "العمر",
        "رقم جوال الطوارئ",
        "رقم التسجيل",
        "رقم جوال ولي الأمر",
        "إسم ولي الأمر"
    ]
    
    # Write headers for children
    for col, header in enumerate(children_headers, start=1):
        cell = ws_children.cell(row=1, column=col)
        cell.value = header
        cell.font = cell.font.copy(bold=True)
        cell.alignment = cell.alignment.copy(horizontal="center")
    
    # Get all children
    children = Children.query.all()
    
    # Write children data
    for row, child in enumerate(children, start=2):
        # Get parent's data
        parent = Registration.query.filter_by(phone_number=child.parent_phone).first()
        parent_full_name = f"{parent.first_name} {parent.father_name} {parent.first_grand_name} {parent.family_name}" if parent else "غير معروف"
        
        data = [
            child.first_name,
            child.father_name,
            child.grandfather_name,
            child.family_name,
            child.gender,
            child.age,
            child.emergency_phone,
            child.registration_number,
            child.parent_phone,
            parent_full_name
        ]
        
        for col, value in enumerate(data, start=1):
            cell = ws_children.cell(row=row, column=col)
            cell.value = value
            cell.alignment = cell.alignment.copy(horizontal="center")
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"registrations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )


# Route to download database backup
@app.route("/backup_db")
@role_required(['admin'])
def backup_db():
    try:
        # Ensure database exists
        if not os.path.exists(db_path):
            flash("قاعدة البيانات غير موجودة", "error")
            return redirect(url_for("admin"))
            
        # Create log directory
        log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Log')
        os.makedirs(log_dir, exist_ok=True)
        
        # Generate timestamp for filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_name = f"registrations_{timestamp}.db"
        
        # Log download
        with open(os.path.join(log_dir, 'backup.log'), 'a', encoding='utf-8') as f:
            f.write(f"{datetime.now()}: Database downloaded as {download_name}\n")
            
        # Return file as attachment
        return send_file(
            db_path,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/x-sqlite3'
        )
            
    except Exception as e:
        flash(f"حدث خطأ أثناء تحميل قاعدة البيانات: {str(e)}", "error")
        # Log error
        with open(os.path.join(log_dir, 'backup.log'), 'a', encoding='utf-8') as f:
            f.write(f"{datetime.now()}: Error during download - {str(e)}\n")
        return redirect(url_for("admin"))

@app.route('/upload/', methods=['GET', 'POST'])
@role_required(['admin'])
def upload_file():
    if request.method == 'POST':
        # Create log directory first
        log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Log')
        os.makedirs(log_dir, exist_ok=True)
        
        try:
            if 'file' not in request.files:
                flash("لم يتم تحديد ملف", "error")
                return redirect(url_for("admin"))
                
            file = request.files['file']
            if file.filename == '':
                flash("لم يتم تحديد ملف", "error")
                return redirect(url_for("admin"))
                
            # Validate file extension
            if not file.filename.endswith('.db'):
                flash("يجب أن يكون الملف بامتداد .db", "error")
                return redirect(url_for("admin"))
                
            # Create backup of current database before restore
            if os.path.exists(db_path):
                backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backup')
                os.makedirs(backup_dir, exist_ok=True)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                pre_restore_backup = os.path.join(backup_dir, f"pre_restore_{timestamp}.db")
                shutil.copy2(db_path, pre_restore_backup)
                
            # Ensure instance directory exists
            os.makedirs(os.path.dirname(db_path), exist_ok=True)
                
            # Save uploaded file
            file.save(db_path)
            
            # Log restore operation
            with open(os.path.join(log_dir, 'restore.log'), 'a', encoding='utf-8') as f:
                f.write(f"{datetime.now()}: Database restored from uploaded file. Previous DB backed up at {pre_restore_backup if os.path.exists(db_path) else 'No previous DB'}\n")
                
            flash("تم إستعادة قاعدة البيانات بنجاح", "success")
        except Exception as e:
            flash(f"حدث خطأ أثناء إستعادة قاعدة البيانات: {str(e)}", "error")
            # Log error
            with open(os.path.join(log_dir, 'restore.log'), 'a', encoding='utf-8') as f:
                f.write(f"{datetime.now()}: Error during restore - {str(e)}\n")
    return redirect(url_for("admin"))

# Route to restore the database from backup
@app.route("/restore_db", methods=["POST"])
@role_required(['admin'])
def restore_db():
    try:
        if 'backup_file' not in request.files:
            flash('لم يتم اختيار ملف', 'error')
            return redirect(url_for('admin'))
            
        file = request.files['backup_file']
        if file.filename == '':
            flash('لم يتم اختيار ملف', 'error')
            return redirect(url_for('admin'))
            
        if file:
            # Create a temporary backup of current database
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if os.environ.get('RENDER'):
                temp_backup = f'/var/data/temp_backup_{timestamp}.db'
            else:
                temp_backup = f'backup/temp_backup_{timestamp}.db'
                
            shutil.copy2(db_path, temp_backup)
            
            try:
                # Restore the uploaded backup
                file.save(db_path)
                flash('تم استعادة قاعدة البيانات بنجاح', 'success')
            except Exception as e:
                # If restore fails, recover from temporary backup
                shutil.copy2(temp_backup, db_path)
                flash(f'فشل في استعادة قاعدة البيانات: {str(e)}', 'error')
            finally:
                # Clean up temporary backup
                if os.path.exists(temp_backup):
                    os.remove(temp_backup)
                    
    except Exception as e:
        flash(f'حدث خطأ: {str(e)}', 'error')
        
    return redirect(url_for('admin'))


# ------------------------------------------------------------------------------
# ''' Prizes Routes '''
# ------------------------------------------------------------------------------

# Create route for displaying a list of prizes
@app.route("/prizes", methods=["GET", "POST"])
def list_prizes():
    prizes = Prize.query.all()
    filters = FilterForm()

    if request.method == "POST":
        allowed_families = filters.family_name.data
        allowed_age_range = filters.age.data
        allowed_gender = filters.gender.data
        prize_id = request.form.get("prize_id")
        print("prize_id:",prize_id)
        
        # Set filter on the next prize
        prize = db.session.get(Prize, prize_id)  # Updated to use session.get()
        if prize:
            prize.allowed_families = ",".join(allowed_families) if allowed_families else None
            prize.allowed_age_range = ",".join(allowed_age_range) if allowed_age_range else None
            prize.allowed_gender = ",".join(allowed_gender) if allowed_gender else None

            # Set is_next to the selected prize only
            for p in prizes:
                p.is_next = p.id == int(prize_id)

            db.session.commit()
            flash("تم تحديد الهدية للسحب بنجاح!", "success")
            
        else:
            flash("الهدية غير موجودة!", "error")

    return render_template("prizes.html", prizes=prizes, filters=filters)


# Create route for adding a new prize
@app.route("/prizes/add", methods=["GET", "POST"])
def add_prize():
    form = PrizeForm()
    if request.method == "POST":
        name = request.form["name"]
        description = request.form["description"]
        allowed_families = request.form["allowed_families"]
        allowed_age_range = request.form["allowed_age_range"]
        allowed_gender = request.form["allowed_gender"]

        prize = Prize(
            name=name,
            description=description,
            allowed_families=allowed_families,
            allowed_age_range=allowed_age_range,
            allowed_gender=allowed_gender,
        )
        db.session.add(prize)
        db.session.commit()

        flash("تم إضافة الهدية بنجاح", "success")
        return redirect(url_for("list_prizes"))

    return render_template("add_prize.html", form=form)


# Create route for editing a prize
@app.route("/prizes/edit/<int:id>", methods=["GET", "POST"])
def edit_prize(id):
    prize = Prize.query.get(id)
    form = PrizeForm(obj=prize)

    if request.method == "POST":
        prize.name = request.form["name"]
        prize.description = request.form["description"]
        prize.allowed_families = request.form["allowed_families"]
        prize.allowed_age_range = request.form["allowed_age_range"]
        prize.allowed_gender = request.form["allowed_gender"]

        db.session.commit()

        flash("تم تعديل البيانات بنجاح", "success")
        return redirect(url_for("list_prizes"))

    return render_template("edit_prize.html", form=form, prize=prize)


# Create route for deleting a prize
@app.route("/prizes/delete/<int:id>")
@role_required(['admin'])
def delete_prize(id):
    try:
        prize = Prize.query.get(id)
        if prize:
            # If prize is assigned to a guest, reset the guest's prize_id
            if prize.guest_registration_number:
                guest = Registration.query.filter_by(registration_number=prize.guest_registration_number).first()
                if guest:
                    guest.prize_id = None
                    db.session.add(guest)
            
            # Delete the prize
            db.session.delete(prize)
            db.session.commit()
            flash("تم حذف الهدية بنجاح", "success")
        else:
            flash("لم يتم العثور على الهدية", "error")
        return redirect(url_for("list_prizes"))
    except Exception as e:
        db.session.rollback()
        flash(f"حدث خطأ أثناء حذف الهدية: {str(e)}", "error")
        return redirect(url_for("list_prizes"))


# Create route to delete registration number from the prize and search for registration
# number in 'Registeration' table and delete 'prize_id'
@app.route("/prizes/reset/<int:id>", methods=["POST"])
@role_required(['admin'])
def reset_prize(id):
    try:
        # Get the prize
        prize = Prize.query.get_or_404(id)
        if not prize.guest_registration_number:
            flash("لا يوجد رقم تسجيل مرتبط بهذه الجائزة", "warning")
            return redirect(url_for("list_prizes"))
            
        # Find the registration with this number
        sel_reg = Registration.query.filter_by(registration_number=prize.guest_registration_number).first()
        
        # Update prize
        prize.guest_registration_number = None
        
        # Update registration if it exists
        if sel_reg:
            sel_reg.prize_id = None
            
        # Commit changes
        db.session.commit()
        
        # Log the reset
        log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Log')
        os.makedirs(log_dir, exist_ok=True)
        with open(os.path.join(log_dir, 'prize_reset.log'), 'a', encoding='utf-8') as f:
            f.write(f"{datetime.now()}: Prize {prize.name} (ID: {prize.id}) reset. Previous registration number: {prize.guest_registration_number}\n")
            
        flash("تم حذف رقم التسجيل بنجاح!", "success")
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error resetting prize {id}: {str(e)}")
        flash(f"حدث خطأ أثناء إعادة تعيين الجائزة: {str(e)}", "error")
        
    return redirect(url_for("list_prizes"))

# Route to check current prize
@app.route('/check_current_prize')
@role_required(['admin'])
def check_current_prize():
    # Check if user is authenticated
    if 'role' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
        
    try:
        current_prize = Prize.query.filter_by(is_next=True).first()
        if current_prize:
            return jsonify({
                'id': current_prize.id,
                'name': current_prize.name,
                'description': current_prize.description
            })
        return jsonify({'id': None})
    except Exception as e:
        app.logger.error(f"Error checking prize: {str(e)}")
        return jsonify({'error': str(e)}), 500

# ------------------------------------------------------------------------------
# ''' Withdrawal Routes '''
# ------------------------------------------------------------------------------

# Route to withdraw a prize
@app.route("/withdraw_prize", methods=["GET", "POST"])
def withdraw_prize():
    sel_prize = Prize.query.filter_by(
        is_next=True, guest_registration_number=None
    ).first()
    
    print("\n=== Withdrawal Debug Info ===")
    print(f"Selected Prize: {sel_prize.name if sel_prize else 'None'}")
    if sel_prize:
        print(f"Prize ID: {sel_prize.id}")
        print(f"Is Next: {sel_prize.is_next}")
        print(f"Guest Registration Number: {sel_prize.guest_registration_number}")
    
    # Fetch all registration numbers not associated with a prize
    return render_template("withdraw_prize.html", prize=sel_prize)


def get_filtered_reg_no(prize_id: int):
    # Get sel prize filters
    sel_prize = Prize.query.filter_by(id=prize_id).first()
    allowed_families = sel_prize.allowed_families
    allowed_age_range = sel_prize.allowed_age_range
    allowed_gender = sel_prize.allowed_gender

    print("\n=== Withdrawal Filter Debug Info ===")
    print(f"Prize ID: {prize_id}")
    print(f"Allowed Families: {allowed_families}")
    print(f"Allowed Age Range: {allowed_age_range}")
    print(f"Allowed Gender: {allowed_gender}")

    # Base query to filter records without a prize_id
    base_query = Registration.query.filter(
        Registration.prize_id == None, Registration.is_attended == True
    )

    # Build the filter conditions dynamically
    filter_conditions = []

    # Check if allowed_families is not empty
    if allowed_families:
        families = [f.strip() for f in allowed_families.split(",")]
        if "أخرى" in families and "السديس" in families:
            # If both are selected, no family filter needed
            pass
        elif "أخرى" in families:
            # If only "أخرى" is selected, exclude "السديس"
            filter_conditions.append(Registration.family_name != "السديس")
        elif "السديس" in families:
            # If only "السديس" is selected
            filter_conditions.append(Registration.family_name == "السديس")
        else:
            # For any other specific families
            filter_conditions.append(Registration.family_name.in_(families))

    # Check if allowed_age_range is not empty
    if allowed_age_range:
        age_ranges = [age.strip() for age in allowed_age_range.split(",")]
        if "الكل" not in age_ranges:
            age_range_condition = or_(
                *[Registration.age.like(f"%{age_range}%") for age_range in age_ranges]
            )
            filter_conditions.append(age_range_condition)

    # Check if allowed_gender is not empty
    if allowed_gender:
        genders = [gender.strip() for gender in allowed_gender.split(",")]
        if "الكل" not in genders:
            gender_condition = or_(
                *[Registration.gender.like(f"%{gender}%") for gender in genders]
            )
            filter_conditions.append(gender_condition)

    # Apply the filter conditions
    if filter_conditions:
        final_query = base_query.filter(and_(*filter_conditions))
    else:
        final_query = base_query

    print("\n=== Query Debug Info ===")
    print(f"SQL Query: {final_query}")

    # Get the filtered guests
    filtered_guests = final_query.all()
    print("\n=== Filtered Guests ===")
    print(f"Total Filtered Guests: {len(filtered_guests)}")
    
    return filtered_guests


# Route to get random registration number
@app.route("/shuffle_numbers/<int:id>", methods=["GET", "POST"])
def shuffle_numbers(id):
    print("\n=== Starting Shuffle Numbers ===")
    print(f"Prize ID: {id}")
    
    registrations_without_prize = get_filtered_reg_no(prize_id=id)
    print(f"\n=== After Filtering ===")
    print(f"Number of eligible registrations: {len(registrations_without_prize)}")
    
    ic(registrations_without_prize)
    # registrations_without_prize = Registration.query.filter_by(prize_id=None).all()
    random.shuffle(registrations_without_prize)

    # If there are registrations without a prize
    # if registrations_without_prize and len(registrations_without_prize) > 1:
    if registrations_without_prize:
        # Randomly select one registration number
        selected_registration = random.choice(registrations_without_prize)
        print(f"\n=== Selected Winner ===")
        print(f"Name: {selected_registration.first_name} {selected_registration.father_name} {selected_registration.family_name}")
        print(f"Registration Number: {selected_registration.registration_number}")
        print(f"Is Attended: {selected_registration.is_attended}")
        
        ic(selected_registration)
        return jsonify(
            winner=selected_registration.registration_number,
            winner_name=selected_registration.first_name
                    + " "
                    + selected_registration.father_name
                    + " "
                    + selected_registration.first_grand_name
                    + " "
                    + selected_registration.second_grand_name
                    + " "
                    + selected_registration.third_grand_name
                    + " "
                    + selected_registration.family_name,
        )
    else:
        print("\n=== No Eligible Registrations Found ===")
        return jsonify(winner="", winner_name="")


@app.route("/confirm_prize/<int:prize_id>/<int:reg_no>", methods=["GET", "POST"])
def confirm_prize(prize_id, reg_no):
    # Get the sel prize object and update it
    sel_prize = db.session.get(Prize, prize_id)  # Updated to use session.get()
    sel_prize.guest_registration_number = reg_no
    sel_prize.is_next = False
    # Update the registeration table
    db.session.query(Registration).filter(
        Registration.registration_number == reg_no
    ).update({Registration.prize_id: prize_id})
    # Commit the changes
    db.session.commit()
    return jsonify("تم بنجاح")


# ------------------------------------------------------------------------------
# ''' Attendence confirmation Routes '''
# ------------------------------------------------------------------------------

# Route for attendence confirmation
@app.route("/confirm_attendence", methods=["GET", "POST"])
@role_required(['admin', 'receptionist'])
def confirm_attendence():
    form = RegistrationForm()  # Create an instance of the RegistrationForm
    phone = request.args.get('phone')  # Get phone from query params
    if phone:
        form.phone_number.data = phone  # Pre-fill the phone number
    
    if request.method == "POST":
        phone_number = request.form["phone_number"]
        sel_reg = Registration.query.filter_by(phone_number=phone_number).first()
        if sel_reg:
            sel_reg.is_attended = True
            db.session.commit()
            
            # Get children information
            children = Children.query.filter_by(parent_phone=phone_number).all()
            children_names = [child.first_name for child in children]

            # Construct full name
            full_name = f"{sel_reg.first_name} {sel_reg.father_name} {sel_reg.first_grand_name} {sel_reg.family_name}"
            
            return render_template(
                "confirm_attendence.html",
                form=form,
                result="success",
                reg_no=sel_reg.registration_number,
                children_count=len(children),
                children_names=children_names,
                full_name=full_name
            )
        else:
            return render_template(
                "confirm_attendence.html", form=form, result="failed"
            )
    return render_template(
        "confirm_attendence.html", form=form, result=""
    )  # Pass the form instance to the template

@app.route("/confirm_attendance")
def confirm_attendance():
    phone = request.args.get('phone')
    if phone:
        # Pre-fill the phone number in the form
        return render_template("confirm_attendance.html", phone=phone)
    return redirect(url_for("index"))

# ------------------------------------------------------------------------------
# ''' Cards Routes '''
# ------------------------------------------------------------------------------

@app.route("/cards")
def cards():
    return render_template("cards.html")


@app.route("/downloadcard1", methods=["POST"])
def download_card1():
    # print(">>>>>>>>>>>>>>>>>>","Hello")
    text = request.form["text"]
    font = request.form["font"]
    ic(font)
    if not font:
        font = "ReemKufi-Regular.ttf"
    font_path = os.path.join(app.static_folder, "fonts", font)
    ic(font_path)

    image_path = os.path.join(app.static_folder, "card8.jpg")
    img = Image.open(image_path)
    width, height = img.size
    img_draw = ImageDraw.Draw(img)
    img_draw.text(
        xy=(width / 2 - len(text) * 10 , height / 2 + 200),
        text=text,
        font=ImageFont.truetype(font_path, 72),
        fill=(0, 0, 0),
        direction="rtl",
        language="arabic",
    )
    img_io = BytesIO()
    img.save(img_io, "PNG")
    img_io.seek(0)
    return send_file(
        img_io, mimetype="image/png", as_attachment=True, download_name="card.png"
    )


@app.route("/downloadcard2", methods=["POST"])
def download_card2():
    # print(">>>>>>>>>>>>>>>>>>", "Hello")
    text = request.form["text"]
    font = request.form["font"]
    if not font:
        font = "ReemKufi-Regular.ttf"
    font_path = os.path.join(app.static_folder, "fonts", font)

    image_path = os.path.join(app.static_folder, "card7.jpg")
    img = Image.open(image_path)
    width, height = img.size
    img_draw = ImageDraw.Draw(img)
    img_draw.text(
        xy=(width / 2 - len(text) * 10 -150, height / 2 + 300),
        text=text,
        font=ImageFont.truetype(font_path, size=72),
        fill=(0, 0, 0),
        direction="rtl",
        language="arabic",
    )
    img_io = BytesIO()
    img.save(img_io, "PNG")
    img_io.seek(0)
    return send_file(
        img_io, mimetype="image/png", as_attachment=True, download_name="card.png"
    )


@app.route("/ticket/<phone_number>")
def ticket(phone_number):
    registration = Registration.query.filter_by(phone_number=phone_number).first()
    if registration:
        # Get children information
        children = Children.query.filter_by(parent_phone=phone_number).all()
        children_names = [child.first_name for child in children]
        return render_template(
            "ticket.html", 
            registration=registration, 
            children=children,
            children_names=children_names,
            children_count=len(children)
        )
    return redirect(url_for("index"))

@app.route("/verify/<registration_number>")
def verify_ticket(registration_number):
    # Check main registration
    registration = Registration.query.filter_by(registration_number=registration_number).first()
    if registration:
        return jsonify({
            "valid": True,
            "type": "main",
            "name": f"{registration.first_name} {registration.father_name} {registration.family_name}",
            "phone": registration.phone_number
        })
    
    # Check children registration
    child = Children.query.filter_by(registration_number=registration_number).first()
    if child:
        return jsonify({
            "valid": True,
            "type": "child",
            "name": f"{child.first_name} {child.father_name} {child.family_name}",
            "parent_phone": child.parent_phone
        })
    
    return jsonify({"valid": False})

@app.route("/short_ticket/<phone_number>")
def short_ticket(phone_number):
    # Get registration data
    registration = Registration.query.filter_by(phone_number=phone_number).first_or_404()
    
    # Get children if any
    children = None
    if registration.gender == "أنثى":
        children = Children.query.filter_by(parent_phone=phone_number).all()
    
    return render_template("short_ticket.html", 
                         registration=registration, 
                         children=children)

# Route to clear all data from database
@app.route("/clear_all_data")
@role_required(['admin'])
def clear_all_data():
    try:
        # Delete all prizes first to handle foreign key constraints
        Prize.query.delete()
        # Delete all registrations
        Registration.query.delete()
        # Delete all children
        Children.query.delete()
        
        db.session.commit()
        flash("تم مسح جميع البيانات بنجاح", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"حدث خطأ أثناء مسح البيانات: {str(e)}", "error")
    
    return redirect(url_for("admin"))

# ==============================================================================
if __name__ == "__main__":
    app.run(host="0.0.0.0", debug=True)
